import openpyxl
import os


LETTERS = ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K')


def main():
    wb = openpyxl.load_workbook(f'{os.getcwd()}/tab.xlsx')
    sheets = wb.sheetnames
    sheet = wb.active

    wb_1 = openpyxl.Workbook()
    wb_1.create_sheet(title='Первый лист', index=0)
    sheet_1 = wb_1['Первый лист']
    sheet_1.append(['Субъект',
                    'Мужчины в городах до 100 тыщ',
                    'Мужчины в городах от 100 до 500 тыщ',
                    'Мужчины в городах от 500 тыщ до миллиона',
                    'Женщины в городах до 100 тыщ',
                    'Женщины  в городах от 100 до 500 тыщ',
                    'Женщины в городах от 500 тыщ до миллиона'])
    for i in range(30, 1060):
        all_cities = sheet['B' + str(i)].value
        b = i % 11
        men_less_100 = 0
        men_less_500 = 0
        men_less_million = 0
        women_less_100 = 0
        women_less_500 = 0
        women_less_million = 0
        if b == 0:
            # Мужчины
            for letter in LETTERS:
                if sheet[str(letter) + str(i+1)].value == '-':
                    count = 0
                else:
                    count = sheet[str(letter) + str(i+1)].value
                if letter in ('C', 'D', 'E', 'F', 'G'):
                    men_less_100 += int(count)
                if letter in ('H', 'I'):
                    men_less_500 += int(count)
                if letter in ('J', 'K'):
                    men_less_million += int(count)
            # Женщины
            for letter in LETTERS:
                if sheet[str(letter) + str(i+2)].value == '-':
                    count = 0
                else:
                    count = sheet[str(letter) + str(i+2)].value
                if letter in ('C', 'D', 'E', 'F', 'G'):
                    women_less_100 += int(count)
                if letter in ('H', 'I'):
                    women_less_500 += int(count)
                if letter in ('J', 'K'):
                    women_less_million += int(count)
            town = sheet['A' + str(i-5)].value
            if all_cities != '-':
                sheet_1.append([town,
                                str(round(men_less_100 / all_cities, 3)),
                                str(round(men_less_500 / all_cities, 3)),
                                str(round(men_less_million / all_cities, 3)),
                                str(round(women_less_100 / all_cities, 3)),
                                str(round(women_less_500 / all_cities, 3)),
                                str(round(women_less_million / all_cities, 3))
                                ])
    wb_1.save(f'{os.getcwd()}/statistic.xlsx')


if __name__ == '__main__':
    main()
